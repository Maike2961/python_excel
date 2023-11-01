from typing import Optional, Tuple, Union
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

#Setando a aparencia padrao do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()#classe principal
        self.layout_config()
        self.appearence()
        self.todo_sistema()
        
    def layout_config(self):
        self.title("Sistema de Gestão de Cliente")
        self.geometry("700x500")
        
    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=["#000", "#fff"]).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=460)
        
    
    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Gestão de Clientes", font=("Century Gothic bold", 24), text_color="#fff").place(x=190, y=10)
        
        span = ctk.CTkLabel(self, text="Por favor, Preencha todo o formulário", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)
        
        
        fichario = pathlib.Path("Clientes.xlsx")
        if fichario.exists():
            pass
        else:
            fichario=Workbook()
            folha=fichario.active
            folha['A1']="Nome completo"
            folha['B1']="Contato"
            folha['C1']="Idade"
            folha['D1']="Genero"
            folha['E1']="Endereco"
            folha['F1']="Observacoes"

            fichario.save("Clientes.xlsx")

        
        #função dos botôes
        def submit():
            
            #Pagando os dados do entrys
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            address = address_value.get()
            gender = gender_combobox.get()
            obs = obs_entry.get(0.0, END)
            
            if name == "" or contact == "" or age == "" or address == "":
                messagebox.showerror("Sistema", "ERRO!\nPor favor preencha todos os campos!")
            else:
                # pegando os dados e passando para Excel
                fichario = openpyxl.load_workbook("Clientes.xlsx")
                folha = fichario.active

                folha.cell(column=1, row=folha.max_row+1, value=name)
                folha.cell(column=2, row=folha.max_row, value=contact)
                folha.cell(column=3, row=folha.max_row, value=age)
                folha.cell(column=4, row=folha.max_row, value=address)
                folha.cell(column=5, row=folha.max_row, value=gender)
                folha.cell(column=6, row=folha.max_row, value=obs)

                fichario.save(r"Clientes.xlsx")
                messagebox.showinfo("Sistema", "Dados salvos com sucesso")
        
        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            obs_entry.delete(0.0, END)
        
        #text Variables
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()        
        
        
        #entrys
        name_entry = ctk.CTkEntry(self, width=350, textvariable= name_value, font=("Century Gothic bold", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable= contact_value, font=("Century Gothic bold", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable= age_value ,font=("Century Gothic bold", 16), fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200, textvariable= address_value , font=("Century Gothic bold", 16), fg_color="transparent")
        
        #ComboBox
        gender_combobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino"], font=("Century Gothic bold", 14),width=150)
        gender_combobox.set("Masculino")
        
        #Entrada de Obs
        obs_entry = ctk.CTkTextbox(self, width=470, height=150, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")
        
        #Labels
        lb_name = ctk.CTkLabel(self, text="Nome Completo", font=("Certury Gothic bold", 16), text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(self, text="Contato", font=("Certury Gothic bold", 16), text_color=["#000", "#fff"])
        lb_age = ctk.CTkLabel(self, text="Idade", font=("Certury Gothic bold", 16), text_color=["#000", "#fff"])
        lb_gander = ctk.CTkLabel(self, text="Gênero", font=("Certury Gothic bold", 16), text_color=["#000", "#fff"])
        lb_adress = ctk.CTkLabel(self, text="Endereço", font=("Certury Gothic bold", 16), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observações", font=("Certury Gothic bold", 16), text_color=["#000", "#fff"])
        
        #buttons
        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=450)
        btn_submit = ctk.CTkButton(self, text="Limpar dados".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=450)
        
        
        #Posicionando os elementos na janela
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)
        
        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)
        
        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)
        
        lb_gander.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)
        
        lb_adress.place(x=50, y=190)
        address_entry.place(x=50, y=220)
        
        lb_obs.place(x=50, y=280)
        obs_entry.place(x=150, y=280)
        
    
    
    
    def change_apm(self, nova_aparecia):
        ctk.set_appearance_mode(nova_aparecia)


if __name__ == "__main__":
    app = App()
    app.mainloop()